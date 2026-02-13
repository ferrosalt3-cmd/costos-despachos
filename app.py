# app.py
# Costos por tarea + Tablero (S/.) — PRO+
# Persistencia GRATIS con Google Sheets (NO se pierde al reiniciar Streamlit)
# Anti-cuota 429: cache + lee/escribe solo cuando corresponde

import pandas as pd
import streamlit as st
import altair as alt

from datetime import datetime, date, time, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

import gspread
from gspread.exceptions import APIError
from google.oauth2.service_account import Credentials

import time as _time


# ---------------------------
# Constantes
# ---------------------------

TIPOS_TAREA_DEFAULT = [
    "Movimiento interno",
    "Despacho a carro",
    "Recepción de mercadería",
    "Picking",
    "Importacion",
    "Exportacion",
    "Limpieza y orden",
    "Otros",
]

PRIORIDADES = ["Alta", "Media", "Baja"]
TIPO_PERSONAL = ["Operativo", "Administrativo"]

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


# ---------------------------
# Helpers
# ---------------------------

def safe_float(x, default=0.0) -> float:
    try:
        if pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


def safe_bool(x, default=False) -> bool:
    try:
        if isinstance(x, bool):
            return x
        s = str(x).strip().lower()
        if s in ("true", "1", "yes", "y", "si", "sí"):
            return True
        if s in ("false", "0", "no", "n"):
            return False
        return default
    except Exception:
        return default


def combine_date_time(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute, t.second)


def hours_between(start_dt: datetime, end_dt: datetime) -> float:
    """Hours diff >= 0. If end before start, assume crossed midnight (+1 day)."""
    if end_dt < start_dt:
        end_dt = end_dt + timedelta(days=1)
    return max((end_dt - start_dt).total_seconds() / 3600.0, 0.0)


def money(x: float) -> str:
    return f"S/ {x:,.2f}"


def is_saturday(d: date) -> bool:
    return d.weekday() == 5


def ensure_columns(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df


def recalc_cost_hora_personal(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Sueldo_mensual"] = df["Sueldo_mensual"].apply(lambda v: safe_float(v, 0.0))
    df["Horas_mes"] = df["Horas_mes"].apply(lambda v: safe_float(v, 0.0))
    df["Activo"] = df["Activo"].apply(lambda v: safe_bool(v, True))
    df["Costo_hora"] = df.apply(
        lambda r: (safe_float(r["Sueldo_mensual"], 0.0) / safe_float(r["Horas_mes"], 0.0))
        if safe_float(r["Horas_mes"], 0.0) > 0
        else 0.0,
        axis=1,
    )
    return df


def recalc_cost_hora_equipos(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Costo_mensual"] = df["Costo_mensual"].apply(lambda v: safe_float(v, 0.0))
    df["Horas_mes"] = df["Horas_mes"].apply(lambda v: safe_float(v, 0.0))
    df["Activo"] = df["Activo"].apply(lambda v: safe_bool(v, True))
    df["Costo_hora"] = df.apply(
        lambda r: (safe_float(r["Costo_mensual"], 0.0) / safe_float(r["Horas_mes"], 0.0))
        if safe_float(r["Horas_mes"], 0.0) > 0
        else 0.0,
        axis=1,
    )
    return df


def default_data():
    personal = pd.DataFrame(
        [
            {"Nombre": "Giancarlo Luna", "Tipo": "Operativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
            {"Nombre": "Mayra Bejarano", "Tipo": "Operativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
            {"Nombre": "Key Mozombite", "Tipo": "Operativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
            {"Nombre": "Emiliano Quispe", "Tipo": "Operativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
            {"Nombre": "Jose Calvino", "Tipo": "Operativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
            {"Nombre": "Miguel Alarcon", "Tipo": "Operativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
            {"Nombre": "Cleber Reyes", "Tipo": "Administrativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
            {"Nombre": "Jaime Motta", "Tipo": "Administrativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
            {"Nombre": "Jose Lopez", "Tipo": "Administrativo", "Sueldo_mensual": 0.0, "Horas_mes": 208, "Activo": True},
        ]
    )
    personal = ensure_columns(personal, ["Costo_hora"])
    personal = recalc_cost_hora_personal(personal)

    equipos = pd.DataFrame(
        [
            {"Codigo": "Apilador 1", "Tipo": "Apilador", "Costo_mensual": 0.0, "Horas_mes": 180, "Activo": True},
            {"Codigo": "Apilador 3", "Tipo": "Apilador", "Costo_mensual": 0.0, "Horas_mes": 180, "Activo": True},
            {"Codigo": "Apilador 4", "Tipo": "Apilador", "Costo_mensual": 0.0, "Horas_mes": 180, "Activo": True},
            {"Codigo": "Montacarga #05", "Tipo": "Montacarga", "Costo_mensual": 0.0, "Horas_mes": 180, "Activo": True},
            {"Codigo": "Montacarga #07", "Tipo": "Montacarga", "Costo_mensual": 0.0, "Horas_mes": 180, "Activo": True},
            {"Codigo": "Montacarga #08", "Tipo": "Montacarga", "Costo_mensual": 0.0, "Horas_mes": 180, "Activo": True},
            {"Codigo": "Montacarga #11", "Tipo": "Montacarga", "Costo_mensual": 0.0, "Horas_mes": 180, "Activo": True},
            {"Codigo": "Transpaleta 3", "Tipo": "Transpaleta", "Costo_mensual": 0.0, "Horas_mes": 200, "Activo": True},
            {"Codigo": "Transpaleta 4", "Tipo": "Transpaleta", "Costo_mensual": 0.0, "Horas_mes": 200, "Activo": True},
        ]
    )
    equipos = ensure_columns(equipos, ["Costo_hora"])
    equipos = recalc_cost_hora_equipos(equipos)

    tareas = pd.DataFrame(
        columns=[
            "ID",
            "Fecha",
            "Tipo_tarea",
            "Prioridad",
            "Nota",
            "Programada",
            "Estado",
            "Inicio",
            "Fin",
            "Horas",
            "Personal_usado",
            "Equipos_usados",
            "Costo_personal",
            "Costo_equipos",
            "Costo_total",
        ]
    )

    config = {
        "horas_lv_efectivas": 8.0,
        "horas_sab_efectivas": 5.5,
        "inicio_lv": "07:30",
        "fin_lv": "16:45",
        "inicio_sab": "07:30",
        "fin_sab": "13:00",
    }

    return {"personal": personal, "equipos": equipos, "tareas": tareas, "config": config}


# ---------------------------
# Google Sheets (anti 429)
# ---------------------------

@st.cache_resource
def gs_client_cached():
    info = dict(st.secrets["gcp_service_account"])
    if "private_key" in info and "\\n" in info["private_key"]:
        info["private_key"] = info["private_key"].replace("\\n", "\n")
    creds = Credentials.from_service_account_info(info, scopes=SCOPES)
    return gspread.authorize(creds)


@st.cache_resource
def open_sheet_cached():
    return gs_client_cached().open_by_url(st.secrets["GSHEET_URL"])


def _get_ws(sh, title: str):
    try:
        return sh.worksheet(title)
    except Exception:
        return sh.add_worksheet(title=title, rows=2000, cols=50)


def _read_ws_df(sh, title: str) -> pd.DataFrame:
    ws = _get_ws(sh, title)
    values = ws.get_all_values()
    if not values:
        return pd.DataFrame()
    headers = values[0]
    rows = values[1:]
    if not headers:
        return pd.DataFrame()
    return pd.DataFrame(rows, columns=headers)


# ✅ FIX PRINCIPAL: NO usar ws.clear() (reduce 2 llamadas -> 1 llamada)
# + retry anti-429/errores temporales
def _write_ws_df(sh, title: str, df: pd.DataFrame):
    ws = _get_ws(sh, title)
    df2 = df.copy().fillna("")

    # fechas/datetimes a string ISO (para no romper)
    for col in df2.columns:
        if pd.api.types.is_datetime64_any_dtype(df2[col]):
            df2[col] = df2[col].dt.strftime("%Y-%m-%d %H:%M:%S")

    values = [df2.columns.tolist()] + df2.astype(str).values.tolist()

    # 3 reintentos con backoff
    last_err = None
    for attempt in range(3):
        try:
            ws.update("A1", values)  # 1 sola llamada
            return
        except APIError as e:
            last_err = e
            _time.sleep(2 * (attempt + 1))

    # Si sigue fallando, levanta error claro
    raise last_err


@st.cache_data(ttl=25)
def load_store_from_gs():
    sh = open_sheet_cached()

    personal = _read_ws_df(sh, "personal")
    equipos = _read_ws_df(sh, "equipos")
    tareas = _read_ws_df(sh, "tareas")
    config_df = _read_ws_df(sh, "config")

    store = {"personal": personal, "equipos": equipos, "tareas": tareas, "config_df": config_df}
    return store


def save_store_to_gs(store: dict):
    sh = open_sheet_cached()

    _write_ws_df(sh, "personal", store["personal"])
    _write_ws_df(sh, "equipos", store["equipos"])
    _write_ws_df(sh, "tareas", store["tareas"])

    cfg = store["config"]
    cfg_df = pd.DataFrame([{"key": k, "value": v} for k, v in cfg.items()])
    _write_ws_df(sh, "config", cfg_df)

    # para que la siguiente lectura vea lo nuevo sin reventar cuota
    load_store_from_gs.clear()


def normalize_loaded_store(raw: dict) -> dict:
    """Convierte lo leído de Sheets (strings) a tipos correctos y asegura columnas."""
    if raw["personal"].empty and raw["equipos"].empty and raw["tareas"].empty:
        return default_data()

    personal_df = raw["personal"]
    equipos_df = raw["equipos"]
    tareas_df = raw["tareas"]
    config_df = raw["config_df"]

    # Config dict
    config = default_data()["config"]
    if not config_df.empty and "key" in config_df.columns and "value" in config_df.columns:
        for _, r in config_df.iterrows():
            k = str(r.get("key", "")).strip()
            v = r.get("value", "")
            if not k:
                continue
            # intenta parsear numérico si aplica
            if k in ("horas_lv_efectivas", "horas_sab_efectivas"):
                config[k] = safe_float(v, config.get(k, 0.0))
            else:
                config[k] = str(v)

    # Normalizar DF personal
    personal_df = ensure_columns(personal_df, ["Nombre", "Tipo", "Sueldo_mensual", "Horas_mes", "Activo", "Costo_hora"])
    personal_df["Sueldo_mensual"] = personal_df["Sueldo_mensual"].apply(lambda x: safe_float(x, 0.0))
    personal_df["Horas_mes"] = personal_df["Horas_mes"].apply(lambda x: safe_float(x, 208))
    personal_df["Activo"] = personal_df["Activo"].apply(lambda x: safe_bool(x, True))
    personal_df["Tipo"] = personal_df["Tipo"].fillna("Operativo")
    personal_df = recalc_cost_hora_personal(personal_df)

    # Normalizar DF equipos
    equipos_df = ensure_columns(equipos_df, ["Codigo", "Tipo", "Costo_mensual", "Horas_mes", "Activo", "Costo_hora"])
    equipos_df["Costo_mensual"] = equipos_df["Costo_mensual"].apply(lambda x: safe_float(x, 0.0))
    equipos_df["Horas_mes"] = equipos_df["Horas_mes"].apply(lambda x: safe_float(x, 180))
    equipos_df["Activo"] = equipos_df["Activo"].apply(lambda x: safe_bool(x, True))
    equipos_df["Tipo"] = equipos_df["Tipo"].fillna("")
    equipos_df = recalc_cost_hora_equipos(equipos_df)

    # Normalizar DF tareas
    tareas_df = ensure_columns(
        tareas_df,
        [
            "ID","Fecha","Tipo_tarea","Prioridad","Nota","Programada","Estado",
            "Inicio","Fin","Horas","Personal_usado","Equipos_usados",
            "Costo_personal","Costo_equipos","Costo_total",
        ],
    )
    tareas_df["ID"] = pd.to_numeric(tareas_df["ID"], errors="coerce").fillna(0).astype(int)
    tareas_df["Fecha"] = pd.to_datetime(tareas_df["Fecha"], errors="coerce")
    tareas_df["Programada"] = tareas_df["Programada"].apply(lambda x: safe_bool(x, True))
    tareas_df["Horas"] = pd.to_numeric(tareas_df["Horas"], errors="coerce").fillna(0.0)
    for c in ["Costo_personal", "Costo_equipos", "Costo_total"]:
        tareas_df[c] = pd.to_numeric(tareas_df[c], errors="coerce").fillna(0.0)

    # Inicio/Fin parse
    tareas_df["Inicio"] = pd.to_datetime(tareas_df["Inicio"], errors="coerce")
    tareas_df["Fin"] = pd.to_datetime(tareas_df["Fin"], errors="coerce")

    # Estado default
    tareas_df["Estado"] = tareas_df["Estado"].fillna("PROGRAMADA")

    return {"personal": personal_df, "equipos": equipos_df, "tareas": tareas_df, "config": config}


# ---------------------------
# Charts
# ---------------------------

def alt_bar_with_labels(df: pd.DataFrame, x: str, y: str, title: str, horizontal=False):
    if df.empty:
        st.info("Sin datos para este gráfico.")
        return

    base = alt.Chart(df).properties(title=title, height=240)

    if horizontal:
        bars = base.mark_bar().encode(
            y=alt.Y(x, sort="-x", title=None),
            x=alt.X(y, title=None),
            tooltip=[x, y],
        )
        text = base.mark_text(align="left", dx=3).encode(
            y=alt.Y(x, sort="-x"),
            x=alt.X(y),
            text=alt.Text(y, format=",.2f" if df[y].dtype != "int64" else ".0f"),
        )
    else:
        bars = base.mark_bar().encode(
            x=alt.X(x, sort="-y", title=None),
            y=alt.Y(y, title=None),
            tooltip=[x, y],
        )
        text = base.mark_text(dy=-5).encode(
            x=alt.X(x, sort="-y"),
            y=alt.Y(y),
            text=alt.Text(y, format=",.2f" if df[y].dtype != "int64" else ".0f"),
        )

    st.altair_chart((bars + text).interactive(), use_container_width=True)


def export_excel_report(dia: date, tareas_dia: pd.DataFrame, resumen: dict, por_personal: pd.DataFrame, por_tipo: pd.DataFrame) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resumen_df = pd.DataFrame(
            [
                {"Indicador": "Día", "Valor": dia.isoformat()},
                {"Indicador": "Operaciones (total)", "Valor": resumen["ops_total"]},
                {"Indicador": "Finalizadas", "Valor": resumen["finalizadas"]},
                {"Indicador": "Pendientes (no finalizadas)", "Valor": resumen["pendientes"]},
                {"Indicador": "Horas productivas", "Valor": resumen["horas_prod"]},
                {"Indicador": "Horas disponibles", "Valor": resumen["horas_disp"]},
                {"Indicador": "Productividad del turno (%)", "Valor": resumen["prod_pct"]},
                {"Indicador": "Cumplimiento del plan (%)", "Valor": resumen["cumpl_pct"]},
                {"Indicador": "Costo total (finalizadas)", "Valor": resumen["costo_total"]},
            ]
        )
        resumen_df.to_excel(writer, sheet_name="Resumen", index=False)

        detalle = tareas_dia.copy()
        for c in ["Inicio", "Fin", "Fecha"]:
            if c in detalle.columns:
                detalle[c] = detalle[c].astype(str)
        detalle.to_excel(writer, sheet_name="Detalle", index=False)

        por_personal.to_excel(writer, sheet_name="Por_personal", index=False)
        por_tipo.to_excel(writer, sheet_name="Por_tipo", index=False)

    output.seek(0)

    wb = load_workbook(output)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
            max_len = 0
            for v in col:
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)

    if "Por_personal" in wb.sheetnames:
        ws = wb["Por_personal"]
        headers = [c.value for c in ws[1]]
        if "Personal" in headers and "Tareas" in headers and ws.max_row >= 2:
            col_personal = headers.index("Personal") + 1
            col_tareas = headers.index("Tareas") + 1
            chart = BarChart()
            chart.title = "Tareas por colaborador"
            chart.y_axis.title = "Tareas"
            data = Reference(ws, min_col=col_tareas, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=col_personal, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.dLbls = DataLabelList()
            chart.dLbls.showVal = True
            ws.add_chart(chart, "F2")

    out2 = BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2.getvalue()


# ---------------------------
# App
# ---------------------------

st.set_page_config(page_title="Costos Despachos — PRO+", layout="wide")
st.title("📦 Costos por tarea + Tablero (S/.) — PRO+")

# Carga inicial (cacheado)
if "store" not in st.session_state:
    raw = load_store_from_gs()
    st.session_state.store = normalize_loaded_store(raw)

store = st.session_state.store
personal_df = store["personal"]
equipos_df = store["equipos"]
tareas_df = store["tareas"]
config = store["config"]

# Sidebar Configuración (se guarda con botón, para no reventar cuota)
st.sidebar.header("⚙️ Configuración")

cfg_lv = st.sidebar.number_input("Horas disponibles L-V (efectivas)", value=float(config.get("horas_lv_efectivas", 8.0)), step=0.25)
cfg_sab = st.sidebar.number_input("Horas disponibles Sábado (efectivas)", value=float(config.get("horas_sab_efectivas", 5.5)), step=0.25)

st.sidebar.caption("Horario informativo (solo referencia):")
cfg_ini_lv = st.sidebar.text_input("Inicio L-V", value=str(config.get("inicio_lv", "07:30")))
cfg_fin_lv = st.sidebar.text_input("Fin L-V", value=str(config.get("fin_lv", "16:45")))
cfg_ini_sab = st.sidebar.text_input("Inicio Sábado", value=str(config.get("inicio_sab", "07:30")))
cfg_fin_sab = st.sidebar.text_input("Fin Sábado", value=str(config.get("fin_sab", "13:00")))

if st.sidebar.button("💾 Guardar configuración"):
    config["horas_lv_efectivas"] = float(cfg_lv)
    config["horas_sab_efectivas"] = float(cfg_sab)
    config["inicio_lv"] = cfg_ini_lv
    config["fin_lv"] = cfg_fin_lv
    config["inicio_sab"] = cfg_ini_sab
    config["fin_sab"] = cfg_fin_sab

    store["config"] = config
    save_store_to_gs(store)
    st.sidebar.success("Configuración guardada.")
    st.rerun()

st.sidebar.divider()
st.sidebar.caption("Tip: Si hay errores, mira Logs en 'Administrar la aplicación'.")

tabs = st.tabs(["🗓️ Programar", "⏱️ Ejecutar (Iniciar/Finalizar)", "🧑‍🤝‍🧑 Personal", "🚜 Equipos", "📊 Panel de control"])


# ---------------------------
# TAB 1: Programar
# ---------------------------
with tabs[0]:
    st.subheader("🗓️ Programar tareas (plan del día)")
    st.caption("Programa tareas (sin inicio/fin). Luego en Ejecutar las inicias y finalizas para calcular el costo.")

    c1, c2 = st.columns([2, 1])
    with c1:
        tipo_tarea = st.selectbox("Tipo de tarea", TIPOS_TAREA_DEFAULT)
        nota = st.text_input("Nota (opcional)", value="")
    with c2:
        fecha_prog = st.date_input("Fecha programada", value=date.today())
        prioridad = st.selectbox("Prioridad", PRIORIDADES, index=1)

    if st.button("➕ Programar tarea", type="primary"):
        next_id = 1 if tareas_df.empty else int(pd.to_numeric(tareas_df["ID"], errors="coerce").fillna(0).max()) + 1
        new_row = {
            "ID": next_id,
            "Fecha": pd.Timestamp(fecha_prog),
            "Tipo_tarea": tipo_tarea,
            "Prioridad": prioridad,
            "Nota": nota,
            "Programada": True,
            "Estado": "PROGRAMADA",
            "Inicio": pd.NaT,
            "Fin": pd.NaT,
            "Horas": 0.0,
            "Personal_usado": "",
            "Equipos_usados": "",
            "Costo_personal": 0.0,
            "Costo_equipos": 0.0,
            "Costo_total": 0.0,
        }
        tareas_df = pd.concat([tareas_df, pd.DataFrame([new_row])], ignore_index=True)
        store["tareas"] = tareas_df
        save_store_to_gs(store)
        st.success("Tarea programada y guardada en Google Sheets.")
        st.rerun()

    st.divider()
    st.markdown("### 📋 Programadas por día")
    ver_dia = st.date_input("Ver programadas del día", value=date.today(), key="ver_prog_dia")

    df_dia = tareas_df.copy()
    df_dia["Fecha"] = pd.to_datetime(df_dia["Fecha"], errors="coerce")
    df_prog = df_dia[(df_dia["Fecha"].dt.date == ver_dia) & (df_dia["Estado"] == "PROGRAMADA")].copy()

    if df_prog.empty:
        st.info("No hay tareas programadas para ese día.")
    else:
        show_cols = ["ID", "Fecha", "Tipo_tarea", "Prioridad", "Nota", "Estado"]
        st.dataframe(df_prog[show_cols].sort_values("ID"), use_container_width=True)


# ---------------------------
# TAB 2: Ejecutar
# ---------------------------
with tabs[1]:
    st.subheader("⏱️ Ejecutar tareas (PROGRAMADA → EN_CURSO → FINALIZADA)")
    st.caption("Selecciona el día, inicia tareas programadas y luego finalízalas asignando personal/equipos para costear.")

    dia_trabajo = st.date_input("Día de trabajo", value=date.today(), key="dia_trabajo")

    df = tareas_df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df_dia = df[df["Fecha"].dt.date == dia_trabajo].copy()

    colA, colB = st.columns(2)

    with colA:
        st.markdown("#### ✅ PROGRAMADAS (listas para iniciar)")
        prog = df_dia[df_dia["Estado"] == "PROGRAMADA"].copy().sort_values("ID")
        if prog.empty:
            st.info("No hay programadas para hoy.")
        else:
            st.dataframe(prog[["ID", "Tipo_tarea", "Prioridad", "Nota"]], use_container_width=True, height=220)

            ids_prog = prog["ID"].tolist()
            sel_id = st.selectbox("Selecciona ID para iniciar", ids_prog, key="sel_inicio_id")
            default_start = datetime.now().time().replace(second=0, microsecond=0)
            t_inicio = st.time_input("Hora inicio", value=default_start, key="t_inicio")

            if st.button("▶️ Iniciar tarea", key="btn_iniciar"):
                idx = tareas_df.index[tareas_df["ID"] == sel_id]
                if len(idx) == 1:
                    i = idx[0]
                    start_dt = combine_date_time(dia_trabajo, t_inicio)
                    tareas_df.loc[i, "Inicio"] = start_dt
                    tareas_df.loc[i, "Estado"] = "EN_CURSO"
                    tareas_df.loc[i, "Horas"] = 0.0
                    store["tareas"] = tareas_df
                    save_store_to_gs(store)
                    st.success(f"Tarea {sel_id} iniciada y guardada.")
                    st.rerun()

    with colB:
        st.markdown("#### ⏳ EN CURSO (finalizar)")
        en_curso = df_dia[df_dia["Estado"] == "EN_CURSO"].copy().sort_values("ID")

        if en_curso.empty:
            st.info("No hay tareas en curso.")
        else:
            st.dataframe(en_curso[["ID", "Tipo_tarea", "Prioridad", "Inicio"]], use_container_width=True, height=220)

            ids_curso = en_curso["ID"].tolist()
            sel_fin_id = st.selectbox("Selecciona ID en curso", ids_curso, key="sel_fin_id")
            default_end = datetime.now().time().replace(second=0, microsecond=0)
            t_fin = st.time_input("Hora fin", value=default_end, key="t_fin")

            activos_personal = personal_df[(personal_df["Activo"] == True) & (personal_df["Nombre"].notna())].copy()
            activos_equipos = equipos_df[(equipos_df["Activo"] == True) & (equipos_df["Codigo"].notna())].copy()

            pers_sel = st.multiselect("Personal", options=activos_personal["Nombre"].tolist(), default=[], key="pers_sel")
            eq_sel = st.multiselect("Equipos", options=activos_equipos["Codigo"].tolist(), default=[], key="eq_sel")

            if st.button("🏁 Finalizar tarea", type="primary", key="btn_finalizar"):
                idx = tareas_df.index[tareas_df["ID"] == sel_fin_id]
                if len(idx) == 1:
                    i = idx[0]
                    inicio_val = tareas_df.loc[i, "Inicio"]

                    if pd.isna(inicio_val) or inicio_val is None:
                        st.error("Esta tarea no tiene hora de inicio. Iníciala primero.")
                    else:
                        start_dt = pd.to_datetime(inicio_val).to_pydatetime()
                        end_dt = combine_date_time(dia_trabajo, t_fin)
                        horas = hours_between(start_dt, end_dt)

                        personal_cost = 0.0
                        if pers_sel:
                            cost_map = dict(zip(activos_personal["Nombre"], activos_personal["Costo_hora"]))
                            personal_cost = sum(safe_float(cost_map.get(n, 0.0), 0.0) for n in pers_sel) * horas

                        equipos_cost = 0.0
                        if eq_sel:
                            cost_map_e = dict(zip(activos_equipos["Codigo"], activos_equipos["Costo_hora"]))
                            equipos_cost = sum(safe_float(cost_map_e.get(c, 0.0), 0.0) for c in eq_sel) * horas

                        total_cost = personal_cost + equipos_cost

                        tareas_df.loc[i, "Fin"] = end_dt
                        tareas_df.loc[i, "Horas"] = float(round(horas, 2))
                        tareas_df.loc[i, "Personal_usado"] = ", ".join(pers_sel) if pers_sel else ""
                        tareas_df.loc[i, "Equipos_usados"] = ", ".join(eq_sel) if eq_sel else ""
                        tareas_df.loc[i, "Costo_personal"] = float(round(personal_cost, 2))
                        tareas_df.loc[i, "Costo_equipos"] = float(round(equipos_cost, 2))
                        tareas_df.loc[i, "Costo_total"] = float(round(total_cost, 2))
                        tareas_df.loc[i, "Estado"] = "FINALIZADA"

                        store["tareas"] = tareas_df
                        save_store_to_gs(store)
                        st.success(f"Tarea {sel_fin_id} finalizada. Horas: {horas:.2f} | Total: {money(total_cost)}")
                        st.rerun()

    st.divider()
    st.markdown("### 🧾 Registro del día")
    df2 = tareas_df.copy()
    df2["Fecha"] = pd.to_datetime(df2["Fecha"], errors="coerce")
    df_dia_all = df2[df2["Fecha"].dt.date == dia_trabajo].copy().sort_values("ID")

    if df_dia_all.empty:
        st.info("No hay tareas registradas para este día.")
    else:
        show_cols = [
            "ID","Fecha","Tipo_tarea","Prioridad","Programada","Inicio","Fin","Estado",
            "Horas","Personal_usado","Equipos_usados","Costo_personal","Costo_equipos","Costo_total",
        ]
        st.dataframe(df_dia_all[show_cols], use_container_width=True)


# ---------------------------
# TAB 3: Personal
# ---------------------------
with tabs[2]:
    st.subheader("🧑‍🤝‍🧑 Personal (editable)")
    st.caption("Edita personal. Costo/hora = Sueldo_mensual / Horas_mes. Guarda con botón para evitar cuota 429.")

    st.markdown("#### ➕ Agregar personal rápido")
    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        nuevo_nombre = st.text_input("Nombre", value="", key="nuevo_nombre")
    with c2:
        nuevo_tipo = st.selectbox("Tipo", TIPO_PERSONAL, key="nuevo_tipo")
    with c3:
        if st.button("Agregar", key="btn_add_personal"):
            if nuevo_nombre.strip():
                if (personal_df["Nombre"].astype(str).str.strip().str.lower() == nuevo_nombre.strip().lower()).any():
                    st.warning("Ese nombre ya existe.")
                else:
                    add = pd.DataFrame([{
                        "Nombre": nuevo_nombre.strip(),
                        "Tipo": nuevo_tipo,
                        "Sueldo_mensual": 0.0,
                        "Horas_mes": 208,
                        "Activo": True,
                        "Costo_hora": 0.0
                    }])
                    personal_df = pd.concat([personal_df, add], ignore_index=True)
                    personal_df = recalc_cost_hora_personal(personal_df)
                    store["personal"] = personal_df
                    save_store_to_gs(store)
                    st.success("Personal agregado y guardado.")
                    st.rerun()
            else:
                st.error("Escribe un nombre.")

    st.divider()

    edited = st.data_editor(
        personal_df,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config={
            "Nombre": st.column_config.TextColumn("Nombre"),
            "Tipo": st.column_config.SelectboxColumn("Tipo", options=TIPO_PERSONAL),
            "Sueldo_mensual": st.column_config.NumberColumn("Sueldo mensual (S/.)", step=50),
            "Horas_mes": st.column_config.NumberColumn("Horas mes", step=1),
            "Activo": st.column_config.CheckboxColumn("Activo"),
            "Costo_hora": st.column_config.NumberColumn("Costo_hora", disabled=True),
        },
        key="editor_personal",
    )

    if st.button("💾 Guardar cambios de Personal"):
        personal_df = recalc_cost_hora_personal(edited)
        store["personal"] = personal_df
        save_store_to_gs(store)
        st.success("Personal guardado en Google Sheets.")
        st.rerun()

    st.markdown("#### Costo/hora calculado")
    st.dataframe(personal_df[["Nombre","Tipo","Activo","Sueldo_mensual","Horas_mes","Costo_hora"]], use_container_width=True)


# ---------------------------
# TAB 4: Equipos
# ---------------------------
with tabs[3]:
    st.subheader("🚜 Equipos (editable)")
    st.caption("Edita equipos. Costo/hora = Costo_mensual / Horas_mes. Guarda con botón para evitar cuota 429.")

    edited_e = st.data_editor(
        equipos_df,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config={
            "Codigo": st.column_config.TextColumn("Código"),
            "Tipo": st.column_config.TextColumn("Tipo"),
            "Costo_mensual": st.column_config.NumberColumn("Costo mensual (S/.)", step=50),
            "Horas_mes": st.column_config.NumberColumn("Horas mes", step=1),
            "Activo": st.column_config.CheckboxColumn("Activo"),
            "Costo_hora": st.column_config.NumberColumn("Costo_hora", disabled=True),
        },
        key="editor_equipos",
    )

    if st.button("💾 Guardar cambios de Equipos"):
        equipos_df = recalc_cost_hora_equipos(edited_e)
        store["equipos"] = equipos_df
        save_store_to_gs(store)
        st.success("Equipos guardados en Google Sheets.")
        st.rerun()

    st.markdown("#### Costo/hora calculado")
    st.dataframe(equipos_df[["Codigo","Tipo","Activo","Costo_mensual","Horas_mes","Costo_hora"]], use_container_width=True)


# ---------------------------
# TAB 5: Dashboard
# ---------------------------
with tabs[4]:
    st.subheader("📊 Panel de control (gerencial)")
    st.caption("Resumen del día con métricas y gráficos por prioridad, tipo y colaborador.")

    dia = st.date_input("Día a analizar", value=date.today(), key="dia_dashboard")
    operativos_turno = st.number_input("Operativos en turno (para horas disponibles)", min_value=0, value=6, step=1)

    df = tareas_df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df_dia = df[df["Fecha"].dt.date == dia].copy()

    ops_total = int(len(df_dia))
    finalizadas = int((df_dia["Estado"] == "FINALIZADA").sum())
    pendientes = int(ops_total - finalizadas)

    horas_prod = float(pd.to_numeric(df_dia.loc[df_dia["Estado"] == "FINALIZADA", "Horas"], errors="coerce").fillna(0).sum())
    costo_total = float(pd.to_numeric(df_dia.loc[df_dia["Estado"] == "FINALIZADA", "Costo_total"], errors="coerce").fillna(0).sum())

    horas_por_op = float(config["horas_sab_efectivas"]) if is_saturday(dia) else float(config["horas_lv_efectivas"])
    horas_disp = float(operativos_turno) * float(horas_por_op)
    prod_pct = (horas_prod / horas_disp * 100.0) if horas_disp > 0 else 0.0

    programadas = int((df_dia["Programada"] == True).sum()) if "Programada" in df_dia.columns else ops_total
    cumpl_pct = (finalizadas / programadas * 100.0) if programadas > 0 else 0.0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Operaciones (día)", f"{ops_total}")
    k2.metric("Finalizadas", f"{finalizadas}")
    k3.metric("Pendientes", f"{pendientes}")
    k4.metric("Costo (finalizadas)", money(costo_total))
    k5.metric("Horas productivas", f"{horas_prod:.2f} h")

    k6, k7, k8 = st.columns(3)
    k6.metric("Horas disponibles", f"{horas_disp:.2f} h")
    k7.metric("Productividad del turno (%)", f"{prod_pct:.1f}%")
    k8.metric("Cumplimiento del plan (%)", f"{cumpl_pct:.1f}%")

    st.divider()

    g1, g2 = st.columns(2)
    with g1:
        tmp = df_dia.copy()
        tmp["Prioridad"] = tmp["Prioridad"].fillna("Sin prioridad")
        pr = tmp.groupby("Prioridad", as_index=False).size().rename(columns={"size": "Cantidad"})
        alt_bar_with_labels(pr, "Prioridad", "Cantidad", "Operaciones por prioridad", horizontal=False)

    with g2:
        tmp = df_dia[df_dia["Estado"] == "FINALIZADA"].copy()
        tmp["Horas"] = pd.to_numeric(tmp["Horas"], errors="coerce").fillna(0.0)
        tmp["Tipo_tarea"] = tmp["Tipo_tarea"].fillna("Sin tipo")
        ht = tmp.groupby("Tipo_tarea", as_index=False)["Horas"].sum().sort_values("Horas", ascending=False)
        alt_bar_with_labels(ht, "Tipo_tarea", "Horas", "Horas por tipo de tarea", horizontal=True)

    st.divider()

    st.markdown("## ⬇️ Reporte Excel (gerencial)")
    fin2 = df_dia[df_dia["Estado"] == "FINALIZADA"].copy()
    fin2["Horas"] = pd.to_numeric(fin2["Horas"], errors="coerce").fillna(0.0)
    fin2["Costo_total"] = pd.to_numeric(fin2["Costo_total"], errors="coerce").fillna(0.0)

    por_tipo = (
        fin2.groupby("Tipo_tarea", as_index=False)
        .agg(Tareas=("ID", "count"), Horas=("Horas", "sum"), Costo_total=("Costo_total", "sum"))
        .sort_values("Costo_total", ascending=False)
    )

    resumen = {
        "ops_total": ops_total,
        "finalizadas": finalizadas,
        "pendientes": pendientes,
        "horas_prod": float(round(horas_prod, 2)),
        "horas_disp": float(round(horas_disp, 2)),
        "prod_pct": float(round(prod_pct, 1)),
        "cumpl_pct": float(round(cumpl_pct, 1)),
        "costo_total": float(round(costo_total, 2)),
    }

    xlsx_bytes = export_excel_report(
        dia=dia,
        tareas_dia=df_dia.sort_values("ID"),
        resumen=resumen,
        por_personal=pd.DataFrame(columns=["Personal", "Tareas", "Horas", "Costo_total"]),
        por_tipo=por_tipo,
    )

    st.download_button(
        "📥 Descargar Excel (reporte del día)",
        data=xlsx_bytes,
        file_name=f"reporte_{dia.isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
